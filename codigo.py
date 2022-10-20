import csv
import datetime
import sys
import openpyxl
from openpyxl.styles import Font, Alignment

clientes = {}
reservaciones = {}
salas = {}

try:
  with open("reservaciones.csv","r", newline="") as archivo:
      lector = csv.reader(archivo)
      next(lector)
      for clave_reservacion, clave_sala, fecha_reservacion, clave_cliente, nombre_reservacion, clave_turno in lector:
        fecha_reservacion = datetime.datetime.strptime(fecha_reservacion, "%Y-%m-%d").date()
        reservaciones[int(clave_reservacion)] = int(clave_sala), fecha_reservacion, int(clave_cliente), nombre_reservacion, int(clave_turno)

  with open("clientes.csv","r", newline="") as archivo:
      lector = csv.reader(archivo)
      next(lector)
      for clave_cliente, nombre in lector:
          clientes[int(clave_cliente)] = nombre

  with open("salas.csv","r", newline="") as archivo:
      lector = csv.reader(archivo)
      next(lector)
      for clave_sala, nombre, capacidad in lector:
          salas[int(clave_sala)] = nombre, int(capacidad)
except FileNotFoundError:
  print("NO SE ENCUENTRAN ARCHIVOS COMPLETOS, INICIANDO ESTADO DESDE CERO")
except Exception:
  print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
else:
  print("ESTADO ANTERIOR RECUPERADO")

turnos = {1: "MATUTINO", 2: "VESPERTINO", 3: "NOCTURNO"}

while True:
  print(f'\n{"-" * 70}')
  print(f'{" " * 23}    MENU DE OPCIONES    {" " * 23}')
  print(f'{"-" * 70}')
  print(f"\n [1] Reservaciones\n",
        "[2] Reportes\n",
        "[3] Registrar una Sala\n",
        "[4] Registrar Nuevo Cliente\n",
        "[5] Salir")
  print(f"{'~' * 70}")

  while True:
    try:
      opcion_menu = int(input("SELECCIONE UNA OPCION: "))
    except ValueError:
      print("INGRESE UN VALOR NUMERICO\n")
    except Exception:
      print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
    else:
      if opcion_menu in [1,2,3,4,5]:
        break
      else:
        print("INGRESE UNA RESPUESTA VALIDA\n")

  if opcion_menu == 1:
    while True:
      print(f'\n{"-" * 70}')
      print(f'{" " * 23}    SUBMENU DE RESERVACIONES    {" " * 23}')
      print(f'{"-" * 70}')
      print(f"\n [1] Registrar una Reservacion\n",
            "[2] Editar el Nombre de un Evento ya Existente\n",
            "[3] Consultar Disponibilidad de Salas para una Fecha\n")
      print(f"{'~' * 70}")

      while True:
        try:
          opcion_reservaciones = int(input("SELECCIONE UNA OPCION: "))
        except ValueError:
          print("INGRESE UN VALOR NUMERICO\n")
        except Exception:
          print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
        else:
          if opcion_reservaciones in [1,2,3]:
            break
          else:
            print("INGRESE UNA RESPUESTA VALIDA\n")

      if opcion_reservaciones == 1:
        if clientes and salas:
          print(f'\n{"_" * 70}')
          print(f'{" " * 20}  MENU REGISTRO DE UNA RESERVACION  {" " * 20}')
          print(f'{"_" * 70}')

          print("\n****** CLIENTES EXISTENTES ******")
          print("*" * 28)
          print("{:<6} {:<20}".format('CLAVE','NOMBRE'))
          print("*" * 28)
          for clave_cliente, nombre_cliente in clientes.items():
            print("{:<6} {:<20}".format(clave_cliente, nombre_cliente))       
          print(f'{"*" * 28}\n')

          clave_clientes = [clave for clave in clientes.keys()]

          while True:
            try:
              clave_cliente = int(input("INGRESE LA CLAVE DEL CLIENTE: \n"))
            except ValueError:
              print("INGRESE UN VALOR NUMERICO\n")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
            else:
              if clave_cliente in clave_clientes:
                break
              else:
                print("INGRESE UNA CLAVE DE CLIENTE EXISTENTE\n")

          while True:
            try:
              fecha_capturada = input("INGRESE LA FECHA DEL EVENTO: (DD/MM/AAAA) \n")
              fecha_reservacion = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
            except ValueError:
              print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
            else:
              fecha_actual = datetime.date.today()
              fecha_valida = fecha_reservacion - datetime.timedelta(days=2)

              if fecha_actual <= fecha_valida:
                break
              else:
                print("LA RESERVACION DEBE DE HACERSE CON 2 DIAS DE ANTICPACION ANTES DEL EVENTO\n")

          clave_salas = [clave for clave in salas.keys()]

          print("\n****** SALAS EXISTENTES ******")
          print("*" * 28)
          print("{:<6} {:<20}".format('CLAVE','NOMBRE'))
          print("*" * 28)
          for clave_sala, [nombre_sala, capacidad_sala] in salas.items():
            print("{:<6} {:<20}".format(clave_sala, nombre_sala))
          print(f'{"*" * 28}\n')

          while True:
            try:
              clave_sala = int(input("INGRESE LA CLAVE DE LA SALA: \n"))
            except ValueError:
              print("INGRESE UN VALOR NUMERICO")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
            else:
              if clave_sala in clave_salas:
                break
              else:
                print("SE TIENE QUE ELEGIR UNA SALA EXISTENTE\n")

          print(f'{"~" * 70}')
          print(f'{" " * 20}   TURNOS DE RESERVACIONES   {" " * 20}')
          print(f" 1) MATUTINO\n",
                  "2) VESPERTINO\n",
                  "3) NOCTURNO\n")

          while True:
            try:
              clave_turno = int(input("SELECCIONE EL TURNO DE LA RESERVACION: "))
            except ValueError:
              print("INGRESE UN VALOR NUMERICO")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
            else:
              if clave_turno in [1,2,3]:
                break
              else:
                print("INGRESE UNA OPCION VALIDA\n")

          coincidencias = 0
          for sala, fecha, cliente, nombre, turno_reservacion in reservaciones.values():
            if (clave_sala == sala) and (fecha_reservacion == fecha) and (clave_turno == turno_reservacion):
              coincidencias += 1

          if coincidencias == 0:
            while True:
              try:
                nombre_reservacion = input("INGRESE EL NOMBRE DEL EVENTO: ")
              except Exception:
                print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
              else:
                if nombre_reservacion == "":
                  print("EL DATO NO SE PUEDE OMITIR\n")
                else:
                  break

            clave_reservacion = max(list(reservaciones.keys()), default = 0) + 1

            reservaciones[clave_reservacion] = [clave_sala, fecha_reservacion, clave_cliente, nombre_reservacion, clave_turno]

            print(f'{"*" * 10} RESERVACION REGISTRADA CON EXITO {"*" * 10}')

          else:
            print("YA EXISTE UNA RESERVACION EN LA FECHA INGRESADA\n")
            break
          break
        else:
          print("NO SE CUENTA CON CLIENTES O SALAS REGISTRADAS")
          break

      elif opcion_reservaciones == 2:
        print(f'\n{"_" * 70}')
        print(f'{" " * 20}  EDITAR NOMBRE DEL EVENTO  {" " * 20}\n')
        print(f'{"_" * 70}')

        print("\n****** EVENTOS EXISTENTES ******")
        print("*" * 28)
        print("{:<6} {:<20}".format('CLAVE','NOMBRE'))
        print("*" * 28)
        for clave_reservacion, datos in reservaciones.items():
          print("{:<6} {:<20}".format(clave_reservacion, datos[3]))
        print(f'{"*" * 28}\n')

        while True:
          try:
            clave = int(input("INGRESE LA CLAVE DE LA RESERVACION: \n"))
          except ValueError:
            print("INGRESE UN VALOR NUMERICO")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
          else:
            if clave in reservaciones.keys():
              break
            else:
              print("INGRESE UNA CLAVE DE RESERVACION EXISTENTE\n")

        while True:
          try:
            nombre_nuevo = input("INGRESE EL NUEVO NOMBRE DE LA SALA: \n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
          else:
            if nombre_nuevo == "":
              print("EL DATO NO SE PUEDE OMITIR\n")
            else:
              break

        for clave_registrada, [sala, fecha, cliente, nombre, turno_reservacion] in reservaciones.items():
          if clave == clave_registrada:
            reservaciones.update({clave_registrada: [sala, fecha, cliente, nombre_nuevo, turno_reservacion]})

        print(f"***** MODIFICACION EXITOSA *****")
        break

      elif opcion_reservaciones == 3:
        while True:
          try:
            fecha_capturada = input("INGRESE LA FECHA A CONSULTAR: (DD/MM/AAAA) \n")
            fecha = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
          except ValueError:
            print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            break

        print("*" * 28)
        print("DISPONIBILIDAD")
        print("*" * 28)
        print("{:<6} {:<10} {:<20}".format('CLAVE','SALA','TURNO'))
        print("*" * 28)

        reservaciones_hechas = []
        reservaciones_posibles = []

        if reservaciones:
          for clave_reservacion, [clave_sala, fecha_reservacion, clave_clientes, nombre_reservacion, clave_turno] in reservaciones.items():
            if fecha == fecha_reservacion:
              for clave, [nombre_sala, capacidad] in salas.items():
                if clave_sala == clave:
                  reservaciones_hechas.append((clave, nombre_sala, turnos[clave_turno]))
          reservaciones_registradas = set(reservaciones_hechas)

          for clave_sala, [nombre, capacidad] in salas.items():
            for clave_turno, descripcion in turnos.items():
              reservaciones_posibles.append((clave_sala, nombre, descripcion))
          reservaciones_disponibles = set(reservaciones_posibles)
              
          turnos_disponibles = sorted(list(reservaciones_disponibles - reservaciones_registradas))

          for clave_sala, sala, turno in turnos_disponibles:
            print("{:<6} {:<10} {:<20}".format(clave_sala, sala, turno))

        else:
          for clave_sala, [nombre, capacidad] in salas.items():
            for clave_turno, descripcion in turnos.items():
              reservaciones_posibles.append((clave_sala, nombre, descripcion))

          for clave_sala, sala, turno in reservaciones_posibles:
            print("{:<6} {:<10} {:<20}".format(clave_sala, sala, turno))

        print("*" * 28)
        break

  elif opcion_menu == 2:
    while True:
      print(f'\n{"-" * 70}')
      print(f'{" " * 23}    SUBMENU DE REPORTES    {" " * 23}')
      print(f'{"-" * 70}')
      print(f"\n [1] Reporte en Pantalla de Reservaciones para una Fecha\n",
            "[2] Exportar Reporte en Excel\n")
      print(f"{'~' * 70}")

      while True:
        try:
          opcion_reportes = int(input("SELECCIONE UNA OPCION: "))
        except ValueError:
          print("INGRESE UN VALOR NUMERICO")
        except Exception:
          print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
        else:
          if opcion_reportes in [1,2]:
            break
          else:
            print("INGRESE UNA OPCION VALIDA\n")

      if opcion_reportes == 1:
        print(f'\n{"_" * 70}')
        print(f'{" " * 20}  REPORTE DE RESERVACIONES  {" " * 20}')
        print(f'{"_" * 70}')

        while True:
          try:
            fecha_capturada = input("INGRESE LA FECHA DEL EVENTO A CONSULTAR  (DD/MM/AAAA): \n")
            fecha_modificada = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
          except ValueError:
            print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            break

        print("\n" + "*"*77)
        print("**" + " "*13 + f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_capturada}" + " "*13 + "**")
        print("*"*77)
        # Se le asigna la cantidad de espacio disponible para cada columna y se muestra lo que estará en cada espacio
        print("{:<6} {:<20} {:<38} {:<13}".format('SALA','CLIENTE','EVENTO', 'TURNO'))
        print("*" * 77)

        for clave_registrada, [clave_sala, fecha, clave_cliente, nombre, clave_turno] in reservaciones.items():
          if fecha_modificada == fecha:
            print("{:<6} {:<20} {:<38} {:<13}".format(salas[clave_sala][0], clientes[clave_cliente], nombre, turnos[clave_turno]))
        print("*"*30 + " FIN DEL REPORTE " + "*"*30)
        break

      elif opcion_reportes == 2:
        print(f'\n{"_" * 70}')
        print(f'{" " * 20}  REPORTE DE RESERVACIONES EN EXCEL {" " * 20}')
        print(f'{"_" * 70}')

        while True:
          try:
            fecha_capturada = input("INGRESE LA FECHA DEL EVENTO A CONSULTAR  (DD/MM/AAAA): ")
            fecha_modificada = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
          except ValueError:
            print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            break

        libro = openpyxl.Workbook()
        libro.iso_dates = True
        hoja = libro["Sheet"] 
        hoja.title = "Reservaciones"

        hoja.sheet_properties.tabColor = 'A4C5DF'

        hoja["A1"].value = f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_capturada}"

        hoja.merge_cells('A1:H1')
        hoja.merge_cells('A2:B2')
        hoja.merge_cells('C2:D2')
        hoja.merge_cells('E2:F2')
        hoja.merge_cells('G2:H2')

        hoja['A1'].font = Font(bold=True, italic=True, size = 14)
        hoja['A1'].alignment = Alignment (horizontal ='center')
        hoja['A2'].alignment = Alignment (horizontal='center')
        hoja['C2'].alignment = Alignment (horizontal='center')
        hoja['E2'].alignment = Alignment (horizontal='center')
        hoja['G2'].alignment = Alignment (horizontal='center')

        hoja["A2"].value = f"Sala"
        hoja["C2"].value = f"Cliente"
        hoja["E2"].value = f"Nombre"
        hoja["G2"].value = f"Turno"

        renglon = 3
        for clave_registrada, [clave_sala, fecha, clave_cliente, nombre, clave_turno] in reservaciones.items():
          hoja.merge_cells(start_row = renglon, start_column = 1, end_row = renglon, end_column = 2)
          hoja.merge_cells(start_row = renglon, start_column = 3, end_row = renglon, end_column = 4)
          hoja.merge_cells(start_row = renglon, start_column = 5, end_row = renglon, end_column = 6)
          hoja.merge_cells(start_row = renglon, start_column = 7, end_row = renglon, end_column = 8)

          hoja.cell(row = renglon, column = 1).value = salas[clave_sala][0]
          hoja.cell(row = renglon, column = 3).value = clientes[clave_cliente]
          hoja.cell(row = renglon, column = 5).value = nombre
          hoja.cell(row = renglon, column = 7).value = turnos[clave_turno]
          renglon += 1

        libro.save("reporte_excel.xlsx")
        print(f"***** REPORTE CREADO *****")
        break

  elif opcion_menu == 3:
    print(f'\n{"_" * 70}')
    print(f'{" " * 20}  REGISTRO DE UNA SALA  {" " * 20}')
    print(f'{"_" * 70}')

    while True:
      try:
        nombre_sala = input("INGRESE EL NOMBRE DE LA SALA A REGISTRAR: ")
      except Exception:
        print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
      else:
        if nombre_sala == "":
          print("EL DATO NO SE PUEDE OMITIR\n")
        else:
          break

    while True:
      try:
        capacidad = int(input("\nINGRESE LA CAPACIDAD DE LA SALA: "))
      except ValueError:
        print("INGRESE UN VALOR NUMERICO")
      except Exception:
        print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
      else:
        if capacidad <= 0:
          print("LA CAPACIDAD TIENE QUE SER MAYOR QUE 0")
        else:
          break

    clave_sala = max(list(salas.keys()), default = 0) + 1

    salas[clave_sala] = [nombre_sala,capacidad]

    print(f'{"*" * 10} SALA REGISTRADA CON EXITO {"*" * 10}')

  elif opcion_menu == 4:
    print(f'\n{"_" * 70}')
    print(f'{" " * 20}  MENU REGISTRO DE UN NUEVO CLIENTE {" " * 20}')
    print(f'{"_" * 70}')

    while True:
      try:
        nombre_cliente = input("INGRESE EL NOMBRE DEL CLIENTE A REGISTRAR: ")
      except Exception:
        print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
      else:
        if nombre_cliente == "":
          print("EL DATO NO SE PUEDE OMITIR\n")
        else:
          break

    clave_cliente = max(list(clientes.keys()), default = 0) + 1
    
    clientes[clave_cliente] = nombre_cliente

    print(f'\n{"*" * 10} CLIENTE REGISTRADO {"*" * 10}')

  elif opcion_menu == 5:
    with open("reservaciones.csv", "w", newline="") as archivo_reservaciones:
      registro_reservaciones = csv.writer(archivo_reservaciones)
      registro_reservaciones.writerow(("Clave de Reservacion","Clave de Sala","Fecha de Reservacion","Clave del Cliente","Nombre de la Reservacion","Clave del Turno"))
      registro_reservaciones.writerows([(clave_reservacion, clave_sala, fecha_reservacion, clave_cliente, nombre_reservacion, clave_turno) for clave_reservacion, [clave_sala, fecha_reservacion, nombre_cliente, nombre_reservacion, clave_turno] in reservaciones.items()])

    with open("salas.csv", "w", newline="") as archivo_sala:
      registro_sala = csv.writer(archivo_sala)
      registro_sala.writerow(("Clave","Nombre de la Sala","Capacidad"))
      registro_sala.writerows([(clave_sala, nombre, capacidad) for clave_sala, [nombre, capacidad] in salas.items()])

    with open("clientes.csv", "w", newline="") as archivo_clientes:
      registro_clientes = csv.writer(archivo_clientes)
      registro_clientes.writerow(("Clave","Nombre del Cliente"))
      registro_clientes.writerows([(clave_sala, nombre) for clave_sala, nombre in clientes.items()])

    break

print("\n¡GRACIAS POR SU PREFERENCIA!")